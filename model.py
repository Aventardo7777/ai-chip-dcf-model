import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from copy import copy

# =========================================================
# 模型参数设定
# =========================================================
SCENARIOS = ['Bear', 'Base', 'Bull']
START_YEAR = 2025  # 历史基准年
FORECAST_YEARS = [2026, 2027, 2028, 2029, 2030]
FINAL_YEAR = 2030
TAX_RATE = 0.20
WACC_BASE = 0.10
TERMINAL_GROWTH = 0.03

# 情景营收 CAGR 及利润假设
scenario_params = {
    'Bear': {
        'cagr': 0.10,
        'gross_margin': 0.55,
        'opex_pct': 0.20,        # 研发+销售管理占营收比
        'da_pct': 0.08,
        'capex_pct': 0.15,
        'nwc_pct': 0.10,         # 净营运资本变动占营收增量
        'revenue_2025': 20e9      # 200亿
    },
    'Base': {
        'cagr': 0.25,
        'gross_margin': 0.60,
        'opex_pct': 0.18,
        'da_pct': 0.07,
        'capex_pct': 0.12,
        'nwc_pct': 0.08,
        'revenue_2025': 25e9
    },
    'Bull': {
        'cagr': 0.40,
        'gross_margin': 0.65,
        'opex_pct': 0.15,
        'da_pct': 0.06,
        'capex_pct': 0.10,
        'nwc_pct': 0.06,
        'revenue_2025': 30e9
    }
}

# 固定增长率实现 CAGR
def get_growth_rates(cagr):
    # 简单让每年增长率一致
    return [cagr] * 5

# =========================================================
# 创建 Excel 工作簿
# =========================================================
wb = openpyxl.Workbook()
# 样式定义
header_font = Font(bold=True, color='FFFFFF', size=11)
header_fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
light_blue_fill = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')
white_fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
bold_font = Font(bold=True)
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
number_format = '#,##0'
percent_format = '0.00%'
multiplier_format = '0.0x'

def style_header_row(ws, row, max_col):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

def style_data_cell(ws, row, col, fmt=number_format, bold=False):
    cell = ws.cell(row=row, column=col)
    cell.border = thin_border
    cell.alignment = Alignment(horizontal='center')
    cell.number_format = fmt
    if bold:
        cell.font = bold_font

# =========================================================
# 每个情景生成一个 sheet
# =========================================================
for sc in SCENARIOS:
    ws = wb.create_sheet(title=f'{sc}_DCF')
    params = scenario_params[sc]
    rev_2025 = params['revenue_2025']
    growths = get_growth_rates(params['cagr'])
    
    # 构建营收序列
    revenues = [rev_2025]
    for g in growths:
        revenues.append(revenues[-1] * (1 + g))
    # revenues[0]=2025, revenues[1]=2026...
    
    # 写标题行
    headers = ['Item (in $Mn)'] + ['2025A'] + [str(y) for y in FORECAST_YEARS]
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)
    style_header_row(ws, 1, len(headers))
    
    # 辅助写行函数
    r = 2
    def write_row(label, values, fmt=number_format, bold=False):
        nonlocal r
        ws.cell(row=r, column=1, value=label).font = bold_font if bold else Font()
        ws.cell(row=r, column=1).border = thin_border
        for i, v in enumerate(values):
            c = i + 2
            ws.cell(row=r, column=c, value=v)
            style_data_cell(ws, r, c, fmt)
        r += 1
    
    # 营收
    write_row('Revenue', [v/1e6 for v in revenues])
    # 毛利
    gross_margin = params['gross_margin']
    gross_profit = [v * gross_margin / 1e6 for v in revenues]
    write_row('Gross Profit', gross_profit)
    # 运营费用
    opex_pct = params['opex_pct']
    opex = [v * opex_pct / 1e6 for v in revenues]
    write_row('Operating Expenses', opex)
    # EBITDA = GP - OPEX, 这里简化没有其他收入
    ebitda = [gp - op for gp, op in zip(gross_profit, opex)]
    write_row('EBITDA', ebitda)
    # D&A
    da_pct = params['da_pct']
    da = [v * da_pct / 1e6 for v in revenues]
    write_row('Depreciation & Amort.', da)
    # EBIT
    ebit = [eb - d for eb, d in zip(ebitda, da)]
    write_row('EBIT', ebit)
    # 税
    tax = [e * TAX_RATE if e > 0 else 0 for e in ebit]
    write_row('Taxes', tax)
    # NOPAT
    nopat = [e - t for e, t in zip(ebit, tax)]
    write_row('NOPAT', nopat)
    # 加回 D&A
    da_after = da  # 同上
    # CapEx
    capex_pct = params['capex_pct']
    capex = [v * capex_pct / 1e6 for v in revenues]
    write_row('Capital Expenditures', capex)
    # 净营运资本变动 (NWC 增量)
    nwc_pct = params['nwc_pct']
    nwc_balances = [v * nwc_pct / 1e6 for v in revenues]
    delta_nwc = [0]  # 2025 无变化
    for i in range(1, len(nwc_balances)):
        delta_nwc.append(nwc_balances[i] - nwc_balances[i-1])
    write_row('Δ Net Working Capital', delta_nwc)
    # FCFF
    fcff = [nopat[i] + da_after[i] - capex[i] - delta_nwc[i] for i in range(len(nopat))]
    write_row('Free Cash Flow (FCFF)', fcff, bold=True)
    
    # 折现因子 & 现值
    wacc = WACC_BASE
    discount_factors = [1/(1+wacc)**(i) for i in range(len(FORECAST_YEARS))]  # 对应2026为第1年，2030为第5年
    pv_fcff = [fcff[i+1] * discount_factors[i] for i in range(len(FORECAST_YEARS))]  # fcff[1:] 是预测期
    pv_labels = [f'PV of FCFF ({y})' for y in FORECAST_YEARS]
    r_temp = r
    for i, y in enumerate(FORECAST_YEARS):
        ws.cell(row=r_temp+i, column=1, value=f'PV Factor {y}').border = thin_border
        ws.cell(row=r_temp+i, column=2, value=discount_factors[i])
        style_data_cell(ws, r_temp+i, 2, percent_format)
        ws.cell(row=r_temp+i, column=3, value=fcff[i+1])
        style_data_cell(ws, r_temp+i, 3, number_format)
        ws.cell(row=r_temp+i, column=4, value=pv_fcff[i])
        style_data_cell(ws, r_temp+i, 4, number_format)
    r = r_temp + len(FORECAST_YEARS)
    
    # 终值计算
    terminal_value = fcff[-1] * (1 + TERMINAL_GROWTH) / (wacc - TERMINAL_GROWTH)
    pv_terminal = terminal_value / (1 + wacc) ** 5
    ws.cell(row=r, column=1, value='Terminal Value (Gordon)').font = bold_font
    ws.cell(row=r, column=1).border = thin_border
    ws.cell(row=r, column=2, value=terminal_value)
    style_data_cell(ws, r, 2, number_format, bold=True)
    ws.cell(row=r, column=3, value=TERMINAL_GROWTH)
    style_data_cell(ws, r, 3, percent_format)
    ws.cell(row=r, column=4, value=pv_terminal)
    style_data_cell(ws, r, 4, number_format, bold=True)
    r += 1
    
    # 企业价值
    enterprise_value = sum(pv_fcff) + pv_terminal
    ws.cell(row=r, column=1, value='Enterprise Value').font = bold_font
    ws.cell(row=r, column=1).border = thin_border
    ws.cell(row=r, column=4, value=enterprise_value)
    style_data_cell(ws, r, 4, number_format, bold=True)
    # 假设无净债务
    ws.cell(row=r+1, column=1, value='(-) Net Debt').border = thin_border
    ws.cell(row=r+1, column=4, value=0)
    style_data_cell(ws, r+1, 4, number_format)
    ws.cell(row=r+2, column=1, value='Equity Value').font = bold_font
    ws.cell(row=r+2, column=1).border = thin_border
    ws.cell(row=r+2, column=4, value=enterprise_value)
    style_data_cell(ws, r+2, 4, number_format, bold=True)
    
    # 调整列宽
    ws.column_dimensions['A'].width = 28
    for col_letter in ['B', 'C', 'D', 'E', 'F', 'G']:
        ws.column_dimensions[col_letter].width = 18

# =========================================================
# 敏感性分析 Sheet
# =========================================================
ws_sens = wb.create_sheet('Sensitivity_Analysis')
ws_sens.title = 'Sensitivity'
# 使用 Base 情景参数计算敏感性
params = scenario_params['Base']
rev_2025 = params['revenue_2025']
growths = get_growth_rates(params['cagr'])
revenues = [rev_2025]
for g in growths:
    revenues.append(revenues[-1] * (1 + g))
gross_profit = [r * params['gross_margin'] for r in revenues]
opex = [r * params['opex_pct'] for r in revenues]
ebitda = [gp - op for gp, op in zip(gross_profit, opex)]
da = [r * params['da_pct'] for r in revenues]
ebit = [eb - d for eb, d in zip(ebitda, da)]
nopat = [e * (1 - TAX_RATE) for e in ebit]
capex = [r * params['capex_pct'] for r in revenues]
nwc_balances = [r * params['nwc_pct'] for r in revenues]
delta_nwc = [0] + [nwc_balances[i] - nwc_balances[i-1] for i in range(1, len(nwc_balances))]
fcff = [nopat[i] + da[i] - capex[i] - delta_nwc[i] for i in range(len(revenues))]
base_fcff_last = fcff[-1]

def compute_ev(wacc, g):
    pv_sum = 0
    for i in range(1, 6):
        pv_sum += fcff[i] / (1 + wacc) ** i
    terminal = fcff[5] * (1 + g) / (wacc - g)
    pv_term = terminal / (1 + wacc) ** 5
    return pv_sum + pv_term

wacc_values = [0.08, 0.09, 0.10, 0.11, 0.12, 0.13, 0.14]
g_values = [0.01, 0.02, 0.03, 0.04]

# 写表头
ws_sens.cell(row=1, column=1, value='Sensitivity of Enterprise Value (Base Case)')
ws_sens.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(g_values)+1)
ws_sens.cell(row=1, column=1).font = Font(bold=True, size=14, color='2F5496')
ws_sens.cell(row=2, column=1, value='WACC \ g')
for j, g in enumerate(g_values):
    ws_sens.cell(row=2, column=2+j, value=g)
    ws_sens.cell(row=2, column=2+j).number_format = percent_format
    style_data_cell(ws_sens, 2, 2+j, percent_format)
style_header_row(ws_sens, 2, len(g_values)+1)

for i, wacc in enumerate(wacc_values):
    row = 3 + i
    ws_sens.cell(row=row, column=1, value=wacc)
    style_data_cell(ws_sens, row, 1, percent_format)
    for j, g in enumerate(g_values):
        ev = compute_ev(wacc, g)
        ws_sens.cell(row=row, column=2+j, value=ev / 1e6)  # 百万单位
        style_data_cell(ws_sens, row, 2+j, number_format)

# 条件格式：使用色阶（手动设置最低最高中间色）
from openpyxl.formatting.rule import ColorScaleRule
min_col = 2
max_col = len(g_values) + 1
min_row = 3
max_row = 3 + len(wacc_values) - 1
ws_sens.conditional_formatting.add(
    f'{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{max_row}',
    ColorScaleRule(start_type='min', start_color='F8696B',
                   mid_type='percentile', mid_value=50, mid_color='FFEB84',
                   end_type='max', end_color='63BE7B')
)
ws_sens.column_dimensions['A'].width = 12
for col_letter in ['B', 'C', 'D', 'E']:
    ws_sens.column_dimensions[col_letter].width = 16

# 删除默认 sheet
if 'Sheet' in wb.sheetnames:
    del wb['Sheet']

# 保存
output_file = 'AI_Chip_DCF_Model.xlsx'
wb.save(output_file)
print(f'Excel saved as {output_file}')